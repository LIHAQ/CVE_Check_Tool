import argparse
import requests
from prettytable import PrettyTable
from datetime import datetime
import pandas as pd
import openpyxl

def query_cpe_info(product, version):
    # 构造查询CPE名称和版本号的URL
    cpe_query = f'https://services.nvd.nist.gov/rest/json/cpes/2.0?cpeMatchString=cpe:2.3:*:*:{product}:{version}:*:*:*'

    # 发送GET请求并获取响应
    response = requests.get(cpe_query)

    # 检查请求是否成功
    if response.status_code == 200:
        # 解析响应的JSON数据
        data = response.json()

        # 检查结果是否为空
        if data['totalResults'] != 0:
            # 遍历产品列表，找到第一个 deprecated 为 False 的 CPE
            for product in data['products']:
                cpe = product['cpe']
                if not cpe['deprecated']:
                    first_cpe_name = cpe['cpeName']
                    print()
                    print("已获取到CPE名称，正在查询CVE编号")
                    return first_cpe_name

            print()
            print("未找到非弃用的CPE名称和版本号。")
        else:
            print()
            print("未找到匹配的CPE名称和版本号。")
    else:
        print()
        print("CPE查询请求失败，请检查网络连接或URL。")


def query_cve_info(cpe_name):
    # 构造查询CVE的URL
    cve_query = f'https://services.nvd.nist.gov/rest/json/cves/2.0?virtualMatchString={cpe_name}'

    # 发送GET请求并获取响应
    response = requests.get(cve_query)

    # 检查CVE查询的请求是否成功
    if response.status_code == 200:
        # 解析CVE查询的JSON数据
        data = response.json()

        if data['totalResults'] != 0:
            # 定义表头名称
            header_names = ['No.', 'Vuln ID', 'CVSS 2.0/3.0 Severity', 'Published', 'Details']

            # 创建PrettyTable对象，并设置表头对齐为居中
            table = PrettyTable(header_names)
            table.align = 'c'

            # 创建DataFrame对象
            df = pd.DataFrame(columns=header_names)

            # 遍历漏洞列表，获取每个 CVE ID
            for idx, cve_item in enumerate(data['vulnerabilities']):
                # 获取漏洞 ID
                cve_id = cve_item['cve']['id']

                # 获取 CVSS V3.0 评级
                cvss_v3_severity = None
                if 'cvssMetricV30' in cve_item['cve']['metrics']:
                    cvss_v3_severity = cve_item['cve']['metrics']['cvssMetricV30'][0]['cvssData']['baseSeverity']
                elif 'cvssMetricV31' in cve_item['cve']['metrics']:
                    cvss_v3_severity = cve_item['cve']['metrics']['cvssMetricV31'][0]['cvssData']['baseSeverity']

                # 获取 CVSS V2.0 评级
                cvss_v2_severity = None
                if 'cvssMetricV2' in cve_item['cve']['metrics']:
                    cvss_v2_severity = cve_item['cve']['metrics']['cvssMetricV2'][0]['baseSeverity']

                # 如果找不到CVSS评级，则标记为 N/A
                if cvss_v3_severity is None:
                    cvss_v3_severity = "N/A"
                if cvss_v2_severity is None:
                    cvss_v2_severity = "N/A"

                # 获取发布日期并转换为中文时间格式
                published_date = cve_item['cve']['published']
                # 解析ISO 8601格式的时间字符串
                datetime_obj = datetime.fromisoformat(published_date)
                # 格式化为中文时间字符串
                published_date = datetime_obj.strftime("%Y/%m/%d %H:%M")

                # 获取详情页URL
                detail_url = f'https://nvd.nist.gov/vuln/detail/{cve_id}'

                # 添加行数据到表格中
                table.add_row([idx+1, cve_id, f"{cvss_v2_severity} | {cvss_v3_severity}", published_date, detail_url])

                # 添加行数据到DataFrame中
                df = df._append({'No.': idx+1, 'Vuln ID': cve_id, 'CVSS 2.0/3.0 Severity': f"{cvss_v2_severity} | {cvss_v3_severity}", 'Published': published_date, 'Details': detail_url}, ignore_index=True)

            # 打印表格
            print()
            print(f"查询到{idx+1}个CVE，如下表：")
            print()
            print(table)

            return df
        else:
            print()
            print("未找到与该CPE匹配的CVE。")
    else:
        print()
        print("CVE查询请求失败，请检查网络连接或URL。")

# 导出到Excel文件
def export_to_excel(df, output_file):
    # 创建一个新的工作簿
    workbook = openpyxl.Workbook()
    # 创建工作表
    worksheet = workbook.active
    worksheet.title = 'CVE'

    # 设置列宽
    column_widths = [15, 20, 15, 30]
    for i, width in enumerate(column_widths):
        column_letter = openpyxl.utils.get_column_letter(i + 1)
        worksheet.column_dimensions[column_letter].width = width

    # 对单元格应用格式
    align_center = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                    right=openpyxl.styles.Side(style='thin'),
                                    top=openpyxl.styles.Side(style='thin'),
                                    bottom=openpyxl.styles.Side(style='thin'))
    header_font = openpyxl.styles.Font(bold=True)

    # 添加表头
    header_values = list(df.columns)
    worksheet.append(header_values)

    # 对标题行应用格式
    for cell in worksheet[1]:
        cell.alignment = align_center
        cell.border = border
        cell.font = header_font

    # 将数据写入工作表
    for row_index, (_, row) in enumerate(df.iterrows(), start=2):
        for col, value in enumerate(row.values, start=1):
            cell = worksheet.cell(row=row_index, column=col)
            cell.value = value
            cell.alignment = align_center
            cell.border = border

    # 自动调整列宽
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # 自动调整行高
    for row in worksheet.iter_rows():
        max_height = 0
        for cell in row:
            try:
                cell_value = str(cell.value)
                line_count = cell_value.count('\n') + 1
                cell_height = (line_count * 14)
                if cell_height > max_height:
                    max_height = cell_height
            except:
                pass
        worksheet.row_dimensions[row[0].row].height = max_height

    workbook.save(output_file)
    print()
    print(f"CVE信息已导出到 {output_file} 文件。")

def main():
    # 创建命令行参数解析器
    parser = argparse.ArgumentParser(description='查询组件的CVE漏洞信息')
    parser.add_argument('-p', '--product', help='组件名称', required=True)
    parser.add_argument('-v', '--version', help='组件版本', required=True)
    parser.add_argument('-o', '--output', help='导出到Excel文件',const='default_output_file',nargs='?',metavar='<output_file>')

    # 解析命令行参数
    args = parser.parse_args()
    cpe_name = query_cpe_info(args.product, args.version)
    if cpe_name:
        cve_df = query_cve_info(cpe_name)
    if args.output is not None:
      if args.output == 'default_output_file':
        output_file =  f'{args.product}_{args.version}_CVE.xlsx'
        export_to_excel(cve_df, output_file)
      else:
        output_file = args.output
        export_to_excel(cve_df, output_file)
if __name__ == '__main__':
    main()